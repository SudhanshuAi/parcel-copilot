"""Validated ingestion for the immutable ParcelPilot assessment source pack."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from parcelpilot.authority import DOCUMENT_SOURCES, SOURCE_MANIFEST, SOURCE_OVERRIDES, DocumentSource, source_path


class IngestionError(RuntimeError):
    """Raised when a source pack is incomplete, changed, or structurally invalid."""


@dataclass(frozen=True)
class ExtractedChunk:
    section: str
    text: str
    page_number: int
    topics: tuple[str, ...]


EXPECTED_HEADINGS: dict[str, tuple[str, ...]] = {
    "support-policy-v3": ("Scope and source precedence", "Severity definitions", "Default first-response targets"),
    "support-policy-v2": ("DEPRECATED", "Severity and response targets"),
    "cancellation-credit-sop-v4": ("Order cancellation", "Failed-pickup service credits", "Approval and uncertainty"),
    "product-operations-guide": ("Plan capabilities", "KI-208", "KI-211"),
    "northstar-agreement": ("Support terms", "Shipment cancellation", "Service credits"),
    "lumenworks-agreement": ("Support terms", "Cancellation terms", "Failed-pickup credits"),
}

EXPECTED_HEADERS: dict[str, tuple[str, ...]] = {
    "accounts": ("account_id", "account_name", "plan", "status", "csm", "contract_file", "premium_support", "notes"),
    "orders": (
        "order_id", "account_id", "carrier", "status", "booked_at", "pickup_window_start",
        "pickup_window_end", "pickup_actual_at", "shipment_fee_inr", "carrier_fault",
        "customer_fault", "cancellation_requested_at", "notes",
    ),
    "tickets": (
        "ticket_id", "account_id", "created_at", "status", "subject", "description", "channel",
        "assigned_to", "last_customer_message_at", "historical_resolution",
    ),
}

TOPIC_TERMS: dict[str, tuple[str, ...]] = {
    "support_sla": ("response target", "support terms", "p1", "p2", "p3"),
    "severity": ("severity", "p1", "p2", "p3", "outage", "credential exposure"),
    "cancellation": ("cancellation", "cancel", "return-to-origin"),
    "service_credit": ("service credit", "failed-pickup", "pickup is more than", "credit"),
    "product_capability": ("bulk upload", "supported file size", "not included"),
    "known_issue": ("known issues", "ki-", "workaround", "investigating", "monitoring"),
    "shipment_status": ("shipment status", "booked", "picked_up", "webhook"),
}

DURATION_RE = r"\d+\s+(?:business\s+hours?|business\s+days?|minutes?|hours?|days?)(?:,\s*24x7)?"


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def verify_source_manifest(source_dir: Path) -> None:
    failures: list[str] = []
    for file_name, expected_hash in SOURCE_MANIFEST.items():
        try:
            actual_hash = sha256(source_path(source_dir, file_name))
        except FileNotFoundError as exc:
            failures.append(str(exc))
            continue
        if actual_hash != expected_hash:
            failures.append(f"Hash mismatch for {file_name}: expected {expected_hash}, got {actual_hash}")
    if failures:
        raise IngestionError("Source manifest validation failed:\n" + "\n".join(failures))


def extract_pdf_pages(path: Path) -> list[str]:
    reader = PdfReader(str(path))
    pages = [normalize_text(page.extract_text() or "") for page in reader.pages]
    if not pages or not all(pages):
        raise IngestionError(f"PDF text extraction was empty for at least one page: {path.name}")
    return pages


def topics_for_text(text: str, source: DocumentSource) -> tuple[str, ...]:
    lowered = text.lower()
    detected = [topic for topic, terms in TOPIC_TERMS.items() if any(term in lowered for term in terms)]
    return tuple(dict.fromkeys((*detected, *source.topics)))


def heading_chunks(page_text: str, source: DocumentSource, page_number: int) -> list[ExtractedChunk]:
    starts = list(re.finditer(r"(?=\b\d+\.\s+[A-Z])", page_text))
    boundaries = [match.start() for match in starts] or [0]
    boundaries.append(len(page_text))
    chunks: list[ExtractedChunk] = []
    for index, start in enumerate(boundaries[:-1]):
        body = page_text[start:boundaries[index + 1]].strip()
        if not body:
            continue
        heading = normalize_text(body.split("●", 1)[0])
        heading = heading[:160] if heading else "Document content"
        issue_starts = list(re.finditer(r"(?=\bKI-\d+\s+-)", body))
        if issue_starts:
            first_issue = issue_starts[0].start()
            prefix = body[:first_issue].strip()
            if prefix:
                chunks.append(ExtractedChunk(heading, prefix, page_number, topics_for_text(prefix, source)))
            issue_bounds = [match.start() for match in issue_starts] + [len(body)]
            for issue_index, issue_start in enumerate(issue_bounds[:-1]):
                issue_body = body[issue_start:issue_bounds[issue_index + 1]].strip()
                issue_heading = normalize_text(issue_body.split("●", 1)[0])[:160]
                chunks.append(ExtractedChunk(f"{heading} / {issue_heading}", issue_body, page_number, topics_for_text(issue_body, source)))
        else:
            chunks.append(ExtractedChunk(heading, body, page_number, topics_for_text(body, source)))
    return chunks


def extract_document(source_dir: Path, source: DocumentSource) -> tuple[list[ExtractedChunk], str]:
    page_texts = extract_pdf_pages(source_path(source_dir, source.file_name))
    combined = " ".join(page_texts)
    missing = [heading for heading in EXPECTED_HEADINGS[source.source_id] if heading.lower() not in combined.lower()]
    if missing:
        raise IngestionError(f"{source.file_name} did not yield expected headings: {missing}")
    chunks = [chunk for number, text in enumerate(page_texts, start=1) for chunk in heading_chunks(text, source, number)]
    if not chunks:
        raise IngestionError(f"No chunks extracted from {source.file_name}")
    return chunks, combined


def _duration_rows(text: str, plans: tuple[str, ...]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for index, plan in enumerate(plans):
        start_match = re.search(rf"\b{re.escape(plan)}\b", text)
        if not start_match:
            raise IngestionError(f"Could not find {plan} SLA row")
        end = len(text)
        if index + 1 < len(plans):
            next_match = re.search(rf"\b{re.escape(plans[index + 1])}\b", text[start_match.end():])
            if not next_match:
                raise IngestionError(f"Could not find end of {plan} SLA row")
            end = start_match.end() + next_match.start()
        durations = re.findall(DURATION_RE, text[start_match.end():end], flags=re.IGNORECASE)
        if len(durations) != 3:
            raise IngestionError(f"Expected three SLA targets for {plan}, found {durations!r}")
        rows.append({"plan": plan, "p1": durations[0], "p2": durations[1], "p3": durations[2]})
    return rows


def _labeled_sla(text: str) -> dict[str, str]:
    values: dict[str, str] = {}
    for severity in ("P1", "P2", "P3"):
        match = re.search(rf"\b{severity}:\s*({DURATION_RE})", text, flags=re.IGNORECASE)
        if not match:
            raise IngestionError(f"Could not parse {severity} support term")
        values[severity.lower()] = normalize_text(match.group(1))
    return values


def _integer(value: str) -> int:
    return int(value.replace(",", ""))


def extract_typed_rules(source: DocumentSource, text: str) -> list[dict[str, Any]]:
    """Extract auditable typed facts from the text, failing on unexpected source structure."""
    rules: list[dict[str, Any]] = []
    if source.source_id in {"support-policy-v3", "support-policy-v2"}:
        rows = _duration_rows(text, ("Enterprise", "Growth", "Standard"))
        for row in rows:
            rules.append({"rule_key": f"support_sla.{row['plan'].lower()}", "topic": "support_sla", "payload": row})
    elif source.source_id == "cancellation-credit-sop-v4":
        fee_match = re.search(r"No fee within (\d+) minutes.*?charge INR ([\d,]+)", text, flags=re.IGNORECASE)
        credit_match = re.search(
            r"more than (\d+) hours.*?lower of INR ([\d,]+) or (\d+)% of the shipment fee",
            text,
            flags=re.IGNORECASE,
        )
        approval_match = re.search(r"credit above INR ([\d,]+) requires manager approval", text, flags=re.IGNORECASE)
        if not fee_match or not credit_match or not approval_match:
            raise IngestionError("Could not parse a required cancellation/service-credit SOP rule")
        rules.extend((
            {"rule_key": "cancellation.default_fee", "topic": "cancellation", "payload": {"booked_grace_minutes": int(fee_match.group(1)), "fee_inr": _integer(fee_match.group(2))}},
            {"rule_key": "service_credit.default_failed_pickup", "topic": "service_credit", "payload": {"delay_hours": int(credit_match.group(1)), "cap_inr": _integer(credit_match.group(2)), "shipment_fee_percent": int(credit_match.group(3)) / 100}},
            {"rule_key": "service_credit.manager_approval", "topic": "approval", "payload": {"above_inr": _integer(approval_match.group(1))}},
        ))
    elif source.source_id == "product-operations-guide":
        max_match = re.search(r"up to ([\d,]+) rows per CSV", text, flags=re.IGNORECASE)
        ki208_match = re.search(r"KI-208.*?above approximately ([\d,]+).*?below ([\d,]+) rows", text, flags=re.IGNORECASE)
        ki211_match = re.search(r"KI-211.*?up to (\d+) minutes late", text, flags=re.IGNORECASE)
        if not max_match or not ki208_match or not ki211_match:
            raise IngestionError("Could not parse a required product guide rule")
        rules.extend((
            {"rule_key": "product.bulk_upload.max_rows", "topic": "product_capability", "payload": {"max_rows": _integer(max_match.group(1)), "plans": ["Growth", "Enterprise"]}},
            {"rule_key": "known_issue.KI-208", "topic": "known_issue", "payload": {"status": "Investigating", "failure_above_rows": _integer(ki208_match.group(1)), "workaround_below_rows": _integer(ki208_match.group(2))}},
            {"rule_key": "known_issue.KI-211", "topic": "known_issue", "payload": {"status": "Monitoring", "carrier": "SwiftShip", "webhook_delay_minutes": int(ki211_match.group(1))}},
        ))
    elif source.source_id == "northstar-agreement":
        sla = _labeled_sla(text)
        cap_match = re.search(r"capped at INR ([\d,]+)", text, flags=re.IGNORECASE)
        if not cap_match or "no cancellation fee" not in text.lower():
            raise IngestionError("Could not parse a required Northstar agreement rule")
        rules.extend((
            {"rule_key": "support_sla.enterprise", "topic": "support_sla", "payload": {"plan": "Enterprise", **sla, "coverage": "24x7"}},
            {"rule_key": "cancellation.booked_pre_pickup_fee_waiver", "topic": "cancellation", "payload": {"status": "BOOKED", "before_pickup": True, "fee_inr": 0}},
            {"rule_key": "service_credit.monthly_cap", "topic": "service_credit", "payload": {"monthly_cap_inr": _integer(cap_match.group(1))}},
        ))
    elif source.source_id == "lumenworks-agreement":
        sla = _labeled_sla(text)
        credit_match = re.search(r"more than (\d+) hours.*?fixed INR ([\d,]+)", text, flags=re.IGNORECASE)
        if not credit_match or "no special cancellation-fee waiver" not in text.lower():
            raise IngestionError("Could not parse a required LumenWorks agreement rule")
        rules.extend((
            {"rule_key": "support_sla.growth", "topic": "support_sla", "payload": {"plan": "Growth", **sla, "coverage": "business_hours_only"}},
            {"rule_key": "cancellation.no_special_fee_waiver", "topic": "cancellation", "payload": {"use_default_sop": True}},
            {"rule_key": "service_credit.failed_pickup", "topic": "service_credit", "payload": {"delay_hours": int(credit_match.group(1)), "fixed_credit_inr": _integer(credit_match.group(2))}},
        ))
    return rules


SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE document_sources (
    source_id TEXT PRIMARY KEY,
    file_name TEXT NOT NULL UNIQUE,
    title TEXT NOT NULL,
    document_type TEXT NOT NULL,
    authority_class TEXT NOT NULL,
    status TEXT NOT NULL,
    effective_from TEXT,
    effective_to TEXT,
    account_id TEXT,
    topics_json TEXT NOT NULL,
    override_topics_json TEXT NOT NULL,
    context_only INTEGER NOT NULL CHECK (context_only IN (0, 1)),
    content_hash TEXT NOT NULL
);

CREATE TABLE source_overrides (
    source_id TEXT NOT NULL REFERENCES document_sources(source_id),
    overrides_source_id TEXT NOT NULL REFERENCES document_sources(source_id),
    topic TEXT NOT NULL,
    reason TEXT NOT NULL,
    PRIMARY KEY (source_id, overrides_source_id, topic)
);

CREATE TABLE document_chunks (
    chunk_id INTEGER PRIMARY KEY,
    source_id TEXT NOT NULL REFERENCES document_sources(source_id),
    section TEXT NOT NULL,
    text TEXT NOT NULL,
    page_number INTEGER NOT NULL,
    topics_json TEXT NOT NULL,
    chunk_hash TEXT NOT NULL UNIQUE
);

CREATE VIRTUAL TABLE document_chunks_fts USING fts5(
    chunk_id UNINDEXED,
    source_id UNINDEXED,
    text
);

CREATE TABLE policy_rules (
    source_id TEXT NOT NULL REFERENCES document_sources(source_id),
    rule_key TEXT NOT NULL,
    topic TEXT NOT NULL,
    payload_json TEXT NOT NULL,
    PRIMARY KEY (source_id, rule_key)
);

CREATE TABLE known_issues (
    issue_id TEXT PRIMARY KEY,
    source_id TEXT NOT NULL REFERENCES document_sources(source_id),
    status TEXT NOT NULL,
    payload_json TEXT NOT NULL
);

CREATE TABLE dataset_metadata (
    key TEXT PRIMARY KEY,
    value_json TEXT NOT NULL
);

CREATE TABLE accounts (
    account_id TEXT PRIMARY KEY,
    account_name TEXT NOT NULL,
    plan TEXT NOT NULL,
    status TEXT NOT NULL,
    csm TEXT NOT NULL,
    contract_file TEXT,
    premium_support INTEGER NOT NULL CHECK (premium_support IN (0, 1)),
    notes TEXT NOT NULL
);

CREATE TABLE orders (
    order_id TEXT PRIMARY KEY,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    carrier TEXT NOT NULL,
    status TEXT NOT NULL,
    booked_at TEXT NOT NULL,
    pickup_window_start TEXT NOT NULL,
    pickup_window_end TEXT NOT NULL,
    pickup_actual_at TEXT,
    shipment_fee_inr REAL NOT NULL,
    carrier_fault INTEGER NOT NULL CHECK (carrier_fault IN (0, 1)),
    customer_fault INTEGER NOT NULL CHECK (customer_fault IN (0, 1)),
    cancellation_requested_at TEXT,
    notes TEXT NOT NULL
);

CREATE TABLE tickets (
    ticket_id TEXT PRIMARY KEY,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    created_at TEXT NOT NULL,
    status TEXT NOT NULL,
    subject TEXT NOT NULL,
    description TEXT NOT NULL,
    channel TEXT NOT NULL,
    assigned_to TEXT NOT NULL,
    last_customer_message_at TEXT NOT NULL
);

CREATE TABLE historical_ticket_context (
    ticket_id TEXT PRIMARY KEY REFERENCES tickets(ticket_id),
    historical_resolution TEXT NOT NULL,
    context_only INTEGER NOT NULL DEFAULT 1 CHECK (context_only = 1)
);

CREATE TABLE action_proposals (
    proposal_id TEXT PRIMARY KEY,
    action_type TEXT NOT NULL CHECK (action_type = 'escalation'),
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    user_id TEXT NOT NULL,
    session_id TEXT NOT NULL,
    target_type TEXT NOT NULL CHECK (target_type IN ('ticket', 'order', 'general_request')),
    target_id TEXT,
    reason_code TEXT NOT NULL,
    summary TEXT NOT NULL,
    evidence_json TEXT NOT NULL,
    payload_hash TEXT NOT NULL,
    status TEXT NOT NULL CHECK (status IN ('proposed', 'confirmed', 'executing', 'executed', 'cancelled', 'expired', 'failed')),
    created_at TEXT NOT NULL,
    expires_at TEXT NOT NULL,
    confirmed_at TEXT,
    executed_at TEXT
);

CREATE TABLE executed_actions (
    action_id TEXT PRIMARY KEY,
    proposal_id TEXT NOT NULL UNIQUE REFERENCES action_proposals(proposal_id),
    action_type TEXT NOT NULL,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    details_json TEXT NOT NULL,
    created_at TEXT NOT NULL
);

CREATE TABLE audit_events (
    audit_id TEXT PRIMARY KEY,
    event_type TEXT NOT NULL,
    actor_user_id TEXT NOT NULL,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    proposal_id TEXT REFERENCES action_proposals(proposal_id),
    action_id TEXT REFERENCES executed_actions(action_id),
    details_json TEXT NOT NULL,
    created_at TEXT NOT NULL
);
"""


def _reset_schema(connection: sqlite3.Connection) -> None:
    tables = (
        "document_chunks_fts", "audit_events", "executed_actions", "action_proposals", "historical_ticket_context", "tickets", "orders", "accounts",
        "dataset_metadata", "known_issues", "policy_rules", "document_chunks", "source_overrides", "document_sources",
    )
    for table in tables:
        connection.execute(f"DROP TABLE IF EXISTS {table}")
    connection.executescript(SCHEMA)


def _db_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return int(value)
    return value


def _worksheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise IngestionError(f"Workbook is missing required sheet {sheet_name!r}")
        sheet = workbook[sheet_name]
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=50))]
        header = header[:len(EXPECTED_HEADERS[sheet_name])]
        if tuple(header) != EXPECTED_HEADERS[sheet_name]:
            raise IngestionError(f"Unexpected {sheet_name} headers: {header!r}")
        rows: list[dict[str, Any]] = []
        for cells in sheet.iter_rows(min_row=2, max_row=100, max_col=len(header), values_only=True):
            if all(value is None for value in cells):
                if rows:
                    break
                continue
            if any(value is None for value in cells if value is not cells[-1]):
                # Nullable spreadsheet fields are handled below; this branch only documents intent.
                pass
            rows.append({key: _db_value(value) for key, value in zip(header, cells, strict=True)})
        if not rows:
            raise IngestionError(f"Workbook sheet {sheet_name!r} has no data rows")
        return rows
    finally:
        workbook.close()


def _read_dataset_metadata(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "README" not in workbook.sheetnames:
            raise IngestionError("Workbook is missing README sheet")
        sheet = workbook["README"]
        values: dict[str, Any] = {}
        for row in sheet.iter_rows(min_row=1, max_row=20, max_col=2, values_only=True):
            if row[0] is not None:
                values[str(row[0])] = _db_value(row[1])
        snapshot = values.get("Dataset snapshot")
        if not snapshot or "Asia/Kolkata" not in str(snapshot):
            raise IngestionError("README Dataset snapshot is missing the required Asia/Kolkata timezone")
        if values.get("Currency") != "INR":
            raise IngestionError("README Currency must be INR")
        return values
    finally:
        workbook.close()


def _insert_source_data(connection: sqlite3.Connection, source_dir: Path) -> tuple[int, int]:
    chunk_count = 0
    rule_count = 0
    for source in DOCUMENT_SOURCES:
        chunks, text = extract_document(source_dir, source)
        connection.execute(
            """INSERT INTO document_sources VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                source.source_id, source.file_name, source.title, source.document_type, source.authority_class,
                source.status, source.effective_from, source.effective_to, source.account_id,
                json.dumps(source.topics), json.dumps(source.override_topics), int(source.context_only),
                sha256(source_path(source_dir, source.file_name)),
            ),
        )
        for chunk in chunks:
            chunk_hash = hashlib.sha256(f"{source.source_id}|{chunk.page_number}|{chunk.section}|{chunk.text}".encode()).hexdigest()
            cursor = connection.execute(
                """INSERT INTO document_chunks(source_id, section, text, page_number, topics_json, chunk_hash)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (source.source_id, chunk.section, chunk.text, chunk.page_number, json.dumps(chunk.topics), chunk_hash),
            )
            connection.execute(
                "INSERT INTO document_chunks_fts(chunk_id, source_id, text) VALUES (?, ?, ?)",
                (cursor.lastrowid, source.source_id, chunk.text),
            )
            chunk_count += 1
        rules = extract_typed_rules(source, text)
        for rule in rules:
            connection.execute(
                "INSERT INTO policy_rules(source_id, rule_key, topic, payload_json) VALUES (?, ?, ?, ?)",
                (source.source_id, rule["rule_key"], rule["topic"], json.dumps(rule["payload"], sort_keys=True)),
            )
            rule_count += 1
            if rule["rule_key"].startswith("known_issue."):
                issue_id = rule["rule_key"].split(".", 1)[1]
                connection.execute(
                    "INSERT INTO known_issues(issue_id, source_id, status, payload_json) VALUES (?, ?, ?, ?)",
                    (issue_id, source.source_id, rule["payload"].get("status", "Unknown"), json.dumps(rule["payload"], sort_keys=True)),
                )
    for override in SOURCE_OVERRIDES:
        connection.execute(
            "INSERT INTO source_overrides(source_id, overrides_source_id, topic, reason) VALUES (?, ?, ?, ?)",
            (override["source_id"], override["overrides_source_id"], override["topic"], override["reason"]),
        )
    return chunk_count, rule_count


def _insert_workbook_data(connection: sqlite3.Connection, workbook_path: Path) -> dict[str, int]:
    metadata = _read_dataset_metadata(workbook_path)
    for key, value in metadata.items():
        connection.execute("INSERT INTO dataset_metadata(key, value_json) VALUES (?, ?)", (key, json.dumps(value)))
    accounts = _worksheet_rows(workbook_path, "accounts")
    orders = _worksheet_rows(workbook_path, "orders")
    tickets = _worksheet_rows(workbook_path, "tickets")
    contract_files = {source.file_name: source.account_id for source in DOCUMENT_SOURCES if source.account_id}
    for row in accounts:
        contract_file = row["contract_file"] or None
        if contract_file and contract_files.get(contract_file) != row["account_id"]:
            raise IngestionError(f"Account {row['account_id']} has an invalid contract file {contract_file!r}")
        connection.execute(
            """INSERT INTO accounts VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (row["account_id"], row["account_name"], row["plan"], row["status"], row["csm"], contract_file, row["premium_support"], row["notes"]),
        )
    for row in orders:
        required = ("order_id", "account_id", "carrier", "status", "booked_at", "pickup_window_start", "pickup_window_end", "shipment_fee_inr")
        if any(row[field] is None for field in required):
            raise IngestionError(f"Order row is missing required values: {row!r}")
        connection.execute(
            """INSERT INTO orders VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            tuple(row[column] for column in EXPECTED_HEADERS["orders"]),
        )
    for row in tickets:
        historical_resolution = row.pop("historical_resolution")
        required = ("ticket_id", "account_id", "created_at", "status", "subject", "description", "channel", "assigned_to", "last_customer_message_at")
        if any(row[field] is None for field in required):
            raise IngestionError(f"Ticket row is missing required values: {row!r}")
        connection.execute(
            """INSERT INTO tickets VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            tuple(row[column] for column in EXPECTED_HEADERS["tickets"][:-1]),
        )
        if historical_resolution:
            connection.execute(
                "INSERT INTO historical_ticket_context(ticket_id, historical_resolution) VALUES (?, ?)",
                (row["ticket_id"], historical_resolution),
            )
    return {"accounts": len(accounts), "orders": len(orders), "tickets": len(tickets)}


def ingest_all(source_dir: Path, database_path: Path) -> dict[str, int]:
    """Create a validated derived SQLite database from the frozen assessment pack."""
    source_dir = source_dir.resolve()
    database_path.parent.mkdir(parents=True, exist_ok=True)
    verify_source_manifest(source_dir)
    connection = sqlite3.connect(database_path)
    try:
        with connection:
            _reset_schema(connection)
            chunk_count, rule_count = _insert_source_data(connection, source_dir)
            workbook_counts = _insert_workbook_data(connection, source_path(source_dir, "ParcelPilot_Assessment_Data.xlsx"))
        return {"documents": len(DOCUMENT_SOURCES), "document_chunks": chunk_count, "policy_rules": rule_count, **workbook_counts}
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build ParcelPilot's validated local source database.")
    parser.add_argument("--source-dir", type=Path, default=Path.cwd(), help="Directory containing the seven supplied source files.")
    parser.add_argument("--database", type=Path, default=Path("data/parcelpilot.db"), help="Output SQLite database path.")
    args = parser.parse_args(list(argv) if argv is not None else None)
    report = ingest_all(args.source_dir, args.database)
    print(json.dumps({"database": str(args.database), **report}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
